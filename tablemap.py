"""
tablemap.py — переиспользуемый инструмент: привести ПРОИЗВОЛЬНУЮ таблицу к заданному
набору выходных полей.

Вход — сырой 2D-грид (строки × ячейки, как выгружается из xlsx со всеми шапками,
пустыми строками и объединёнными ячейками) плюс список желаемых выходных полей.
Большая модель на сервере определяет:
  - какой столбец какому выходному полю соответствует (по смыслу заголовков и данных);
  - с какой строки начинаются собственно данные (шапки/мусор сверху игнорируются).
Извлечение записей по найденной разметке — чистая функция без ИИ (тестируется отдельно).

Инструмент общий: сегодня им разбираем штатное расписание (ФИО, должность, отдел…),
завтра — любую таблицу под любой набор колонок. Знаний о штатке здесь нет — только
механика «вход-таблица + желаемые поля → нормализованные записи».
"""

import os
import re
import json
from typing import Optional

import requests

from config import OLLAMA_URL

LLM_MODEL = os.environ.get("NEIROMASTER_TABLEMAP_MODEL", "qwen3:14b")
LLM_TIMEOUT = 180
SAMPLE_ROWS = 22   # сколько первых строк показываем модели (чтобы видеть 2+ групповых баннера)


def _cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))   # «1.0» -> «1», табельный/№ из числовых ячеек без хвоста .0
    return str(v).strip()


def _llm(system: str, user: str) -> str:
    r = requests.post(f"{OLLAMA_URL}/api/chat", json={
        "model": LLM_MODEL,
        "messages": [{"role": "system", "content": system},
                     {"role": "user", "content": user}],
        "stream": False, "think": False,
    }, timeout=LLM_TIMEOUT)
    r.raise_for_status()
    return r.json()["message"]["content"].strip()


def _parse_json(text: str) -> dict:
    text = re.sub(r"```json\s*|```", "", text)
    i = text.find("{")
    if i == -1:
        raise ValueError("нет JSON в ответе модели")
    # raw_decode берёт ПЕРВЫЙ валидный объект и игнорирует хвост (модель иногда дописывает
    # второй объект или пояснение -> обычный json.loads падал бы с «Extra data»)
    obj, _ = json.JSONDecoder().raw_decode(text[i:])
    return obj


MAP_SYSTEM = """Ты — парсер табличных выгрузок. Тебе дают первые строки таблицы (с индексами
строк и столбцов, начиная с 0) и список ЖЕЛАЕМЫХ ПОЛЕЙ на выходе. ВНИМАТЕЛЬНО разбери
структуру таблицы, учитывая явные И неявные заголовки. Определи:
1) с какого индекса строки начинаются собственно ДАННЫЕ (шапки, заголовки, пустые строки
   сверху — не данные). Если данные разбиты на группы строками-баннерами (см. п.3),
   data_start_row — это строка ПЕРВОГО баннера/группы, чтобы ни одна группа не потерялась;
2) какой индекс СТОЛБЦА соответствует каждому желаемому полю (по смыслу заголовков и
   образца значений). Если подходящего столбца нет — null.
3) СЕКЦИОННЫЕ (групповые) поля: иногда значение поля задаётся не в каждой строке, а
   строкой-БАННЕРОМ над группой строк. Пример: над блоком сотрудников идёт отдельная
   строка с названием отдела/подразделения («Администрация», «Департамент продаж»,
   «Отдел: …»), затем сотрудники этого отдела, потом следующий баннер и т.д. Строки-баннеры
   НЕ являются данными: в них заполнен только столбец с названием группы, а ключевые поля
   (например ФИО) пустые. Если желаемое поле задаётся так — верни в "sections" ИНДЕКС
   СТОЛБЦА, где стоит значение баннера (оно будет протянуто вниз ко всем строкам группы до
   следующего баннера). Ищи такие неявные группировки внимательно.
Ответ — СТРОГО JSON без пояснений:
{"data_start_row": <int>, "columns": {"<поле>": <индекс столбца или null>, ...},
 "sections": {"<поле>": <индекс столбца баннера>, ...}}
Поля, значение которых берётся из обычного столбца в каждой строке, в "sections" не включай.

ПРИМЕР. Таблица:
строка 8: [0]Подразделение
строка 9: [0]№ | [1]Сотрудник | [6]Табельный номер | [17]Должность | [25]Дата приема
строка 10: [0]Администрация
строка 11: [0]1 | [1]Иванов Иван | [6]00246 | [17]Бухгалтер | [25]06.02.2026
строка 12: [0]2 | [1]Петров Пётр | [6]00110 | [17]Кассир | [25]01.02.2023
строка 13: [0]Отдел продаж
строка 14: [0]1 | [1]Сидоров Сидор | [6]00022 | [17]Менеджер | [25]06.03.2020
Желаемые поля: full_name, position, department, tab_number, start_date.
Правильный ответ:
{"data_start_row": 10, "columns": {"full_name": 1, "position": 17, "department": null, "tab_number": 6, "start_date": 25}, "sections": {"department": 0}}
Почему: строка 9 — ЗАГОЛОВОК (не данные, в записи не попадает). Строки 10 и 13 — БАННЕРЫ
отдела (заполнен только столбец 0, ФИО пустое), поэтому department — это СЕКЦИЯ со столбцом 0,
а НЕ обычный столбец. data_start_row = 10 (первый баннер), чтобы захватить все группы."""


def _grid_preview(grid: list, rows: int = SAMPLE_ROWS) -> str:
    lines = []
    for r, row in enumerate(grid[:rows]):
        cells = [f"[{c}]{_cell(val)}" for c, val in enumerate(row) if _cell(val)]
        lines.append(f"строка {r}: " + " | ".join(cells))
    return "\n".join(lines)


def map_columns(grid: list, target_fields: list) -> dict:
    """Определить разметку таблицы большой моделью.
    target_fields — [{"name": "full_name", "description": "ФИО"}, ...].
    Возвращает {"data_start_row": int, "columns": {name: col_index|None}}."""
    names = [f["name"] for f in target_fields]
    fields_txt = "\n".join(f"- {f['name']}: {f.get('description', '')}" for f in target_fields)
    user = f"Желаемые поля:\n{fields_txt}\n\nПервые строки таблицы:\n{_grid_preview(grid)}"
    try:
        data = _parse_json(_llm(MAP_SYSTEM, user))
    except (ValueError, json.JSONDecodeError):
        # модель ответила прозой — просим ещё раз строго JSON
        data = _parse_json(_llm(MAP_SYSTEM, user + "\n\nВЕРНИ ТОЛЬКО JSON, без слов."))
    cols = data.get("columns") or {}
    secs = data.get("sections") or {}
    # оставляем только запрошенные поля; приводим индексы к int или None
    def _col(v):
        return int(v) if isinstance(v, (int, float)) or (isinstance(v, str) and str(v).isdigit()) else None

    clean, sections = {}, {}
    for name in names:
        clean[name] = _col(cols.get(name))
        sc = _col(secs.get(name))
        if sc is not None:
            sections[name] = sc   # индекс столбца строки-баннера, значение протягивается вниз
    return {"data_start_row": int(data.get("data_start_row") or 0), "columns": clean, "sections": sections}


# Слова-заголовки: если «значение» ключевого поля равно одному из них — это строка шапки,
# а не данные (страховка, если модель ошиблась с data_start_row).
_HEADER_WORDS = {"сотрудник", "фио", "ф.и.о.", "ф. и. о.", "работник", "наименование",
                 "имя", "должность", "№", "n", "no", "п/п", "№ п/п"}


def extract_records(grid: list, mapping: dict, required: Optional[list] = None) -> list:
    """Извлечь записи по найденной разметке — чистая функция без ИИ.
    mapping — результат map_columns. required — поля, без которых строка пропускается
    (по умолчанию первое поле с не-null столбцом). Возвращает список dict по полям."""
    columns = mapping.get("columns") or {}
    sections = mapping.get("sections") or {}
    start = mapping.get("data_start_row") or 0
    if required is None:
        required = [name for name, col in columns.items() if col is not None][:1]

    def cval(row, col):
        return _cell(row[col]) if (col is not None and col < len(row)) else ""

    carried = {}   # секционное поле -> последнее значение строки-баннера (протягивается вниз)
    records = []
    for row in grid[start:]:
        is_data = all(cval(row, columns.get(k)) for k in required) if required else True
        if not is_data:
            # строка-баннер группы: обновляем протягиваемые вниз значения секционных полей
            for name, scol in sections.items():
                v = cval(row, scol)
                if v:
                    carried[name] = v
            continue
        # страховка от попавшей в данные строки-шапки: ключевое поле = слово-заголовок
        if any(cval(row, columns.get(k)).strip().lower() in _HEADER_WORDS for k in required):
            continue
        rec = {name: cval(row, col) for name, col in columns.items()}
        for name in sections:
            if not rec.get(name):
                rec[name] = carried.get(name, "")
        records.append(rec)
    return records


def normalize_table(grid: list, target_fields: list, required: Optional[list] = None) -> dict:
    """Полный проход: разметка моделью + извлечение записей.
    Возвращает {"mapping": {...}, "records": [...]}."""
    mapping = map_columns(grid, target_fields)
    return {"mapping": mapping, "records": extract_records(grid, mapping, required)}


def read_xlsx_grid(source) -> list:
    """Читает xlsx (путь или bytes) в 2D-грид строк активного листа. openpyxl импортируется
    лениво — чистой логике маппинга/извлечения драйвер xlsx не нужен (и её тесты тоже)."""
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(source) if isinstance(source, (bytes, bytearray)) else source,
                                data_only=True, read_only=True)
    ws = wb.active
    return [[_cell(v) for v in row] for row in ws.iter_rows(values_only=True)]


def read_csv_grid(source) -> list:
    """Читает CSV/TSV (путь или bytes) в 2D-грид. Кодировка и разделитель определяются
    автоматически (utf-8/cp1251, ';'/','/tab) — выгрузки из Excel обычно cp1251 с ';'."""
    import io
    import csv
    raw = source if isinstance(source, (bytes, bytearray)) else open(source, "rb").read()
    text = None
    for enc in ("utf-8-sig", "cp1251", "utf-8"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=";,\t")
        delimiter = dialect.delimiter
    except csv.Error:
        head = text[:4096]
        delimiter = ";" if head.count(";") >= head.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [[_cell(c) for c in row] for row in reader]


def read_xls_grid(source) -> list:
    """Читает старый .xls (Excel 97-2003) через xlrd. Даты в .xls хранятся числом —
    конвертируем в дату по datemode книги."""
    import xlrd
    wb = (xlrd.open_workbook(file_contents=bytes(source)) if isinstance(source, (bytes, bytearray))
          else xlrd.open_workbook(source))
    ws = wb.sheet_by_index(0)
    grid = []
    for r in range(ws.nrows):
        row = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                    row.append(dt.strftime("%d.%m.%Y") if (dt.hour == 0 and dt.minute == 0)
                               else dt.strftime("%d.%m.%Y %H:%M"))
                    continue
                except Exception:
                    pass
            row.append(_cell(cell.value))
        grid.append(row)
    return grid


def read_table_grid(source, filename: Optional[str] = None) -> list:
    """Единая точка чтения таблицы по расширению имени файла: CSV/TSV, старый .xls или xlsx."""
    name = (filename or (source if isinstance(source, str) else "")).lower()
    if name.endswith(".csv") or name.endswith(".tsv"):
        return read_csv_grid(source)
    if name.endswith(".xls"):
        return read_xls_grid(source)
    return read_xlsx_grid(source)
